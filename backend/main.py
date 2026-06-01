from fastapi import FastAPI, HTTPException, Query, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
from decimal import Decimal
from openpyxl import load_workbook
from io import BytesIO
import base64
import hashlib
import os
import secrets
import string
import unicodedata

import psycopg2
import psycopg2.errors
import psycopg2.extras


app = FastAPI(title="Sistema de Punteos")

app.add_middleware(
    CORSMiddleware,
    allow_origins=os.getenv("CORS_ORIGINS", "*").split(","),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


DB_CONFIG = {
    "host": os.getenv("DB_HOST"),
    "database": os.getenv("DB_NAME"),
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD"),
    "port": os.getenv("DB_PORT", "5432"),
    "sslmode": os.getenv("DB_SSLMODE", "require"),
}

DB_CLIENT_ENCODING = "UTF8"


def get_connection():
    conn = psycopg2.connect(**DB_CONFIG)
    conn.set_client_encoding(DB_CLIENT_ENCODING)
    return conn


def cerrar_conexion(cursor=None, conn=None):
    try:
        if cursor:
            cursor.close()
    except Exception:
        pass

    try:
        if conn:
            conn.close()
    except Exception:
        pass


def convertir_decimal(valor):
    if isinstance(valor, Decimal):
        return float(valor)
    return valor


def limpiar_fila(fila):
    return {clave: convertir_decimal(valor) for clave, valor in dict(fila).items()}


def normalizar_texto(texto):
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return " ".join(texto.split())


def normalizar_encabezado(texto):
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return texto.replace(" ", "_").replace("-", "_")


def obtener_valor_fila(hoja, fila, indice_columna):
    if indice_columna is None:
        return None
    return hoja.cell(row=fila, column=indice_columna).value


def convertir_numero_excel(valor):
    if valor is None or valor == "":
        return 0
    try:
        return float(valor)
    except Exception:
        raise ValueError(f"Valor numérico inválido: {valor}")


def generar_password(longitud=10):
    alfabeto = string.ascii_uppercase + string.digits
    return "".join(secrets.choice(alfabeto) for _ in range(longitud))


def hash_password(password):
    salt = os.urandom(16)
    iteraciones = 100000
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iteraciones)
    return (
        f"pbkdf2_sha256${iteraciones}$"
        f"{base64.b64encode(salt).decode()}$"
        f"{base64.b64encode(digest).decode()}"
    )


def verificar_password(password, password_hash):
    try:
        algoritmo, iteraciones, salt_b64, digest_b64 = password_hash.split("$")
        if algoritmo != "pbkdf2_sha256":
            return False
        salt = base64.b64decode(salt_b64)
        digest = base64.b64decode(digest_b64)
        nuevo_digest = hashlib.pbkdf2_hmac(
            "sha256",
            password.encode("utf-8"),
            salt,
            int(iteraciones),
        )
        return secrets.compare_digest(nuevo_digest, digest)
    except Exception:
        return False


def validar_punteos(actitudinal, zona, examen):
    valores = [actitudinal, zona, examen]
    if any(valor < 0 for valor in valores):
        raise HTTPException(status_code=400, detail="Los punteos no pueden ser negativos")
    if sum(valores) > 100:
        raise HTTPException(
            status_code=400,
            detail="La suma de actitudinal, zona y examen no puede ser mayor a 100",
        )


def obtener_ids_catalogo(cursor, carrera, grado, curso, ciclo_escolar):
    cursor.execute(
        """
        SELECT
            ca.id_carrera,
            gr.id_grado,
            cu.id_curso,
            ce.id_ciclo
        FROM carreras ca
        CROSS JOIN grados gr
        CROSS JOIN cursos cu
        CROSS JOIN ciclos_escolares ce
        WHERE ca.nombre = %s
          AND gr.nombre = %s
          AND cu.nombre = %s
          AND ce.anio = %s;
        """,
        (carrera, grado, curso, ciclo_escolar),
    )
    catalogo = cursor.fetchone()
    if not catalogo:
        raise HTTPException(
            status_code=404,
            detail="No se encontró carrera, grado, curso o ciclo escolar",
        )
    return catalogo


def validar_curso_docente(cursor, codigo_docente, carrera, grado, curso, ciclo_escolar):
    if not codigo_docente:
        raise HTTPException(status_code=400, detail="Debe enviar el código del docente")

    cursor.execute(
        """
        SELECT dc.id_docente_curso
        FROM docentes_cursos dc
        INNER JOIN docentes d ON d.id_docente = dc.id_docente
        INNER JOIN carreras ca ON ca.id_carrera = dc.id_carrera
        INNER JOIN grados gr ON gr.id_grado = dc.id_grado
        INNER JOIN cursos cu ON cu.id_curso = dc.id_curso
        INNER JOIN ciclos_escolares ce ON ce.id_ciclo = dc.id_ciclo
        WHERE d.codigo_docente = %s
          AND ca.nombre = %s
          AND gr.nombre = %s
          AND cu.nombre = %s
          AND ce.anio = %s
          AND dc.estado = TRUE
          AND d.estado = TRUE;
        """,
        (codigo_docente, carrera, grado, curso, ciclo_escolar),
    )
    if not cursor.fetchone():
        raise HTTPException(
            status_code=403,
            detail="Este docente no tiene asignado ese curso",
        )


def registrar_o_actualizar_nota(
    cursor,
    codigo_carnet,
    carrera,
    grado,
    curso,
    bimestre,
    ciclo_escolar,
    actitudinal,
    zona,
    examen,
    observacion,
    actualizar=True,
):
    query_conflict = """
        ON CONFLICT (id_asignacion, id_curso, id_bimestre)
        DO UPDATE SET
            actitudinal = EXCLUDED.actitudinal,
            zona = EXCLUDED.zona,
            examen = EXCLUDED.examen,
            observacion = EXCLUDED.observacion
    """ if actualizar else ""

    query = f"""
        INSERT INTO notas (
            id_asignacion,
            id_curso,
            id_bimestre,
            actitudinal,
            zona,
            examen,
            observacion
        )
        SELECT
            asi.id_asignacion,
            cu.id_curso,
            bi.id_bimestre,
            %s,
            %s,
            %s,
            %s
        FROM asignaciones asi
        INNER JOIN alumnos a ON asi.id_alumno = a.id_alumno
        INNER JOIN carreras ca ON asi.id_carrera = ca.id_carrera
        INNER JOIN grados gr ON asi.id_grado = gr.id_grado
        INNER JOIN ciclos_escolares ce ON asi.id_ciclo = ce.id_ciclo
        INNER JOIN cursos cu ON cu.nombre = %s
        INNER JOIN bimestres bi ON bi.nombre = %s
        WHERE a.codigo_carnet = %s
          AND ca.nombre = %s
          AND gr.nombre = %s
          AND ce.anio = %s
        {query_conflict}
        RETURNING id_nota;
    """

    cursor.execute(
        query,
        (
            actitudinal,
            zona,
            examen,
            observacion,
            curso,
            bimestre,
            codigo_carnet,
            carrera,
            grado,
            ciclo_escolar,
        ),
    )
    nota = cursor.fetchone()
    if not nota:
        raise HTTPException(
            status_code=404,
            detail="No se encontró alumno, carrera, grado, curso o bimestre",
        )
    return nota


class LoginRequest(BaseModel):
    rol: str
    usuario: str
    password: str


class CambiarPasswordRequest(BaseModel):
    rol: str
    usuario: str
    password_actual: str
    password_nueva: str


class NotaRequest(BaseModel):
    codigo_carnet: str
    carrera: str
    grado: str
    curso: str
    bimestre: str
    ciclo_escolar: int = 2026
    actitudinal: float
    zona: float
    examen: float
    observacion: Optional[str] = None
    codigo_docente: Optional[str] = None


class DocenteRequest(BaseModel):
    codigo_docente: str
    nombre_completo: str
    password: Optional[str] = None


class EstudianteRequest(BaseModel):
    codigo_carnet: str
    nombres: str
    apellidos: str
    carrera: str
    grado: str
    ciclo_escolar: int = 2026
    curso: Optional[str] = None
    password: Optional[str] = None


class CursoRequest(BaseModel):
    nombre: str
    carrera: Optional[str] = None
    grado: Optional[str] = None


class DocenteCursoRequest(BaseModel):
    codigo_docente: str
    carrera: str
    grado: str
    curso: str
    ciclo_escolar: int = 2026


class AsignarCursoEstudiantesRequest(BaseModel):
    codigo_docente: str
    carrera: str
    grado: str
    curso: str
    ciclo_escolar: int = 2026


@app.get("/")
def inicio():
    return {"mensaje": "API del Sistema de Punteos funcionando correctamente"}


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/login")
def login(data: LoginRequest):
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        rol = data.rol.lower().strip()
        usuario = data.usuario.strip()

        cursor.execute(
            """
            SELECT *
            FROM usuarios
            WHERE lower(rol) = %s
              AND usuario = %s
              AND estado = TRUE;
            """,
            (rol, usuario),
        )
        cuenta = cursor.fetchone()

        if not cuenta or not verificar_password(data.password, cuenta["password_hash"]):
            raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")

        if rol == "director":
            return {
                "rol": "director",
                "codigo": cuenta["usuario"],
                "nombre_completo": cuenta["nombre_completo"],
            }

        if rol == "docente":
            cursor.execute(
                """
                SELECT codigo_docente, nombre_completo
                FROM docentes
                WHERE codigo_docente = %s
                  AND estado = TRUE;
                """,
                (usuario,),
            )
            docente = cursor.fetchone()
            if not docente:
                raise HTTPException(status_code=404, detail="Docente no encontrado")
            return {
                "rol": "docente",
                "codigo": docente["codigo_docente"],
                "nombre_completo": docente["nombre_completo"],
            }

        if rol == "estudiante":
            cursor.execute(
                """
                SELECT codigo_carnet, nombres, apellidos
                FROM alumnos
                WHERE codigo_carnet = %s
                  AND estado = TRUE;
                """,
                (usuario,),
            )
            estudiante = cursor.fetchone()
            if not estudiante:
                raise HTTPException(status_code=404, detail="Estudiante no encontrado")
            return {
                "rol": "estudiante",
                "codigo": estudiante["codigo_carnet"],
                "nombres": estudiante["nombres"],
                "apellidos": estudiante["apellidos"],
            }

        raise HTTPException(
            status_code=400,
            detail="Rol inválido. Debe ser director, docente o estudiante",
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.post("/usuarios/cambiar-password")
def cambiar_password(data: CambiarPasswordRequest):
    conn = None
    cursor = None
    try:
        rol = data.rol.lower().strip()
        usuario = data.usuario.strip()
        password_nueva = data.password_nueva.strip()

        if rol not in ("docente", "estudiante"):
            raise HTTPException(status_code=400, detail="Solo docentes y estudiantes pueden cambiar su contraseña")
        if len(password_nueva) < 6:
            raise HTTPException(status_code=400, detail="La nueva contraseña debe tener al menos 6 caracteres")

        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute(
            """
            SELECT rol, usuario, nombre_completo, password_hash
            FROM usuarios
            WHERE lower(rol) = %s
              AND usuario = %s
              AND estado = TRUE;
            """,
            (rol, usuario),
        )
        cuenta = cursor.fetchone()
        if not cuenta or not verificar_password(data.password_actual, cuenta["password_hash"]):
            raise HTTPException(status_code=401, detail="La contraseña actual no es correcta")

        cursor.execute(
            """
            UPDATE usuarios
            SET password_hash = %s,
                password_temporal = %s
            WHERE lower(rol) = %s
              AND usuario = %s
            RETURNING rol, usuario, nombre_completo, password_temporal;
            """,
            (hash_password(password_nueva), password_nueva, rol, usuario),
        )
        cuenta_actualizada = cursor.fetchone()
        conn.commit()
        return {
            "mensaje": "Contraseña actualizada correctamente",
            **limpiar_fila(cuenta_actualizada),
        }

    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.get("/estudiantes/{codigo_carnet}")
def obtener_estudiante(codigo_carnet: str):
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        cursor.execute(
            """
            SELECT *
            FROM vista_promedio_alumnos
            WHERE codigo_carnet = %s
            ORDER BY carrera, grado, curso;
            """,
            (codigo_carnet,),
        )
        resultados = cursor.fetchall()
        if not resultados:
            raise HTTPException(
                status_code=404,
                detail="No se encontró información para este código de carnet",
            )

        estudiante = resultados[0]
        cursos = []
        for fila in resultados:
            cursos.append(
                {
                    "curso": fila["curso"],
                    "primer_bimestre": convertir_decimal(fila["primer_bimestre"]),
                    "segundo_bimestre": convertir_decimal(fila["segundo_bimestre"]),
                    "tercer_bimestre": convertir_decimal(fila["tercer_bimestre"]),
                    "cuarto_bimestre": convertir_decimal(fila["cuarto_bimestre"]),
                    "promedio_final": convertir_decimal(fila["promedio_final"]),
                    "estado": fila["estado"],
                }
            )

        return {
            "codigo_carnet": estudiante["codigo_carnet"],
            "apellidos": estudiante["apellidos"],
            "nombres": estudiante["nombres"],
            "carrera": estudiante["carrera"],
            "grado": estudiante["grado"],
            "ciclo_escolar": estudiante["ciclo_escolar"],
            "cursos": cursos,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.get("/director/punteos")
def director_punteos(
    carrera: Optional[str] = Query(None),
    grado: Optional[str] = Query(None),
    curso: Optional[str] = Query(None),
    codigo_docente: Optional[str] = Query(None),
    ciclo_escolar: Optional[int] = Query(None),
):
    return obtener_punteos(carrera, grado, curso, ciclo_escolar, codigo_docente, None)


@app.get("/docente/punteos")
def docente_punteos(
    codigo_docente: str,
    carrera: Optional[str] = Query(None),
    grado: Optional[str] = Query(None),
    curso: Optional[str] = Query(None),
    ciclo_escolar: Optional[int] = Query(None),
):
    return obtener_punteos(carrera, grado, curso, ciclo_escolar, None, codigo_docente)


def obtener_punteos(carrera, grado, curso, ciclo_escolar, codigo_docente, solo_docente):
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        condiciones = []
        parametros = []

        if carrera:
            condiciones.append("v.carrera = %s")
            parametros.append(carrera)
        if grado:
            condiciones.append("v.grado = %s")
            parametros.append(grado)
        if curso:
            condiciones.append("v.curso = %s")
            parametros.append(curso)
        if ciclo_escolar:
            condiciones.append("v.ciclo_escolar = %s")
            parametros.append(ciclo_escolar)
        if codigo_docente:
            condiciones.append("d.codigo_docente = %s")
            parametros.append(codigo_docente)
        if solo_docente:
            condiciones.append("d.codigo_docente = %s")
            parametros.append(solo_docente)

        where_sql = "WHERE " + " AND ".join(condiciones) if condiciones else ""

        cursor.execute(
            f"""
            SELECT
                v.*,
                d.codigo_docente,
                d.nombre_completo AS docente
            FROM vista_promedio_alumnos v
            LEFT JOIN carreras ca ON ca.nombre = v.carrera
            LEFT JOIN grados gr ON gr.nombre = v.grado
            LEFT JOIN cursos cu ON cu.nombre = v.curso
            LEFT JOIN ciclos_escolares ce ON ce.anio = v.ciclo_escolar
            LEFT JOIN docentes_cursos dc
                   ON dc.id_carrera = ca.id_carrera
                  AND dc.id_grado = gr.id_grado
                  AND dc.id_curso = cu.id_curso
                  AND dc.id_ciclo = ce.id_ciclo
                  AND dc.estado = TRUE
            LEFT JOIN docentes d ON d.id_docente = dc.id_docente
            {where_sql}
            ORDER BY v.carrera, gr.numero, v.apellidos, v.nombres, v.curso;
            """,
            parametros,
        )
        return [limpiar_fila(fila) for fila in cursor.fetchall()]

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.post("/docente/notas")
def insertar_nota(data: NotaRequest):
    return guardar_nota_docente(data, actualizar=False)


@app.put("/docente/notas")
def actualizar_nota(data: NotaRequest):
    return guardar_nota_docente(data, actualizar=True)


def guardar_nota_docente(data: NotaRequest, actualizar: bool):
    conn = None
    cursor = None
    try:
        validar_punteos(data.actitudinal, data.zona, data.examen)
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        validar_curso_docente(
            cursor,
            data.codigo_docente,
            data.carrera,
            data.grado,
            data.curso,
            data.ciclo_escolar,
        )
        nota = registrar_o_actualizar_nota(
            cursor,
            data.codigo_carnet,
            data.carrera,
            data.grado,
            data.curso,
            data.bimestre,
            data.ciclo_escolar,
            data.actitudinal,
            data.zona,
            data.examen,
            data.observacion,
            actualizar=actualizar,
        )
        conn.commit()
        return {
            "mensaje": "Nota guardada correctamente",
            "id_nota": nota["id_nota"],
        }

    except psycopg2.errors.UniqueViolation:
        if conn:
            conn.rollback()
        raise HTTPException(
            status_code=409,
            detail="Ya existe una nota registrada para este alumno, curso y bimestre. Use actualizar.",
        )
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.post("/docente/importar-excel")
async def importar_excel(
    archivo: UploadFile = File(...),
    carrera: str = Form(...),
    grado: str = Form(...),
    curso: str = Form(...),
    bimestre: str = Form(...),
    ciclo_escolar: int = Form(2026),
    codigo_docente: str = Form(...),
):
    conn = None
    cursor = None
    try:
        if not archivo.filename.lower().endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Solo se permiten archivos .xlsx")

        contenido = await archivo.read()
        libro = load_workbook(filename=BytesIO(contenido), data_only=True)
        hoja = libro.active

        encabezados = {}
        for columna in range(1, hoja.max_column + 1):
            encabezados[normalizar_encabezado(hoja.cell(row=1, column=columna).value)] = columna

        col_codigo = encabezados.get("codigo_carnet") or encabezados.get("codigo") or encabezados.get("carnet")
        col_actitudinal = encabezados.get("actitudinal")
        col_zona = encabezados.get("zona")
        col_examen = encabezados.get("examen")
        col_observacion = encabezados.get("observacion")

        if not col_codigo or not col_actitudinal or not col_zona or not col_examen:
            raise HTTPException(
                status_code=400,
                detail="El Excel debe tener las columnas: codigo_carnet, actitudinal, zona y examen",
            )

        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        validar_curso_docente(cursor, codigo_docente, carrera, grado, curso, ciclo_escolar)

        procesados = 0
        errores = []

        for fila in range(2, hoja.max_row + 1):
            codigo_carnet = obtener_valor_fila(hoja, fila, col_codigo)
            if codigo_carnet is None or str(codigo_carnet).strip() == "":
                continue

            codigo_carnet = str(codigo_carnet).strip()
            try:
                actitudinal = convertir_numero_excel(obtener_valor_fila(hoja, fila, col_actitudinal))
                zona = convertir_numero_excel(obtener_valor_fila(hoja, fila, col_zona))
                examen = convertir_numero_excel(obtener_valor_fila(hoja, fila, col_examen))
                observacion = obtener_valor_fila(hoja, fila, col_observacion)
                observacion = "" if observacion is None else str(observacion).strip()
                validar_punteos(actitudinal, zona, examen)

                registrar_o_actualizar_nota(
                    cursor,
                    codigo_carnet,
                    carrera,
                    grado,
                    curso,
                    bimestre,
                    ciclo_escolar,
                    actitudinal,
                    zona,
                    examen,
                    observacion,
                    actualizar=True,
                )
                procesados += 1

            except Exception as e:
                errores.append({"fila": fila, "codigo_carnet": codigo_carnet, "error": str(e)})

        conn.commit()
        return {"mensaje": "Importación finalizada", "procesados": procesados, "errores": errores}

    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error al importar Excel: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.post("/docente/asignar-curso-estudiantes")
def asignar_curso_estudiantes_docente(data: AsignarCursoEstudiantesRequest):
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        validar_curso_docente(
            cursor,
            data.codigo_docente,
            data.carrera,
            data.grado,
            data.curso,
            data.ciclo_escolar,
        )
        catalogo = obtener_ids_catalogo(
            cursor,
            data.carrera,
            data.grado,
            data.curso,
            data.ciclo_escolar,
        )

        cursor.execute(
            """
            INSERT INTO cursos_por_grado (id_carrera, id_grado, id_curso)
            VALUES (%s, %s, %s)
            ON CONFLICT (id_carrera, id_grado, id_curso) DO NOTHING;
            """,
            (catalogo["id_carrera"], catalogo["id_grado"], catalogo["id_curso"]),
        )

        cursor.execute(
            """
            INSERT INTO notas (
                id_asignacion,
                id_curso,
                id_bimestre,
                actitudinal,
                zona,
                examen,
                observacion
            )
            SELECT
                asi.id_asignacion,
                %s,
                b.id_bimestre,
                0,
                0,
                0,
                'Curso asignado por docente'
            FROM asignaciones asi
            CROSS JOIN bimestres b
            INNER JOIN alumnos a ON a.id_alumno = asi.id_alumno
            WHERE asi.id_carrera = %s
              AND asi.id_grado = %s
              AND asi.id_ciclo = %s
              AND a.estado = TRUE
            ON CONFLICT (id_asignacion, id_curso, id_bimestre) DO NOTHING
            RETURNING id_nota;
            """,
            (
                catalogo["id_curso"],
                catalogo["id_carrera"],
                catalogo["id_grado"],
                catalogo["id_ciclo"],
            ),
        )
        notas_creadas = cursor.fetchall()

        cursor.execute(
            """
            SELECT COUNT(*) AS total_estudiantes
            FROM asignaciones asi
            INNER JOIN alumnos a ON a.id_alumno = asi.id_alumno
            WHERE asi.id_carrera = %s
              AND asi.id_grado = %s
              AND asi.id_ciclo = %s
              AND a.estado = TRUE;
            """,
            (
                catalogo["id_carrera"],
                catalogo["id_grado"],
                catalogo["id_ciclo"],
            ),
        )
        total = cursor.fetchone()
        conn.commit()

        return {
            "mensaje": "Curso asignado a estudiantes correctamente",
            "estudiantes": total["total_estudiantes"],
            "notas_creadas": len(notas_creadas),
        }

    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.delete("/docente/asignar-curso-estudiantes")
def quitar_curso_estudiantes_docente(data: AsignarCursoEstudiantesRequest):
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        validar_curso_docente(
            cursor,
            data.codigo_docente,
            data.carrera,
            data.grado,
            data.curso,
            data.ciclo_escolar,
        )
        catalogo = obtener_ids_catalogo(
            cursor,
            data.carrera,
            data.grado,
            data.curso,
            data.ciclo_escolar,
        )

        cursor.execute(
            """
            DELETE FROM notas n
            USING asignaciones asi, alumnos a
            WHERE n.id_asignacion = asi.id_asignacion
              AND asi.id_alumno = a.id_alumno
              AND n.id_curso = %s
              AND asi.id_carrera = %s
              AND asi.id_grado = %s
              AND asi.id_ciclo = %s
              AND a.estado = TRUE
            RETURNING n.id_nota;
            """,
            (
                catalogo["id_curso"],
                catalogo["id_carrera"],
                catalogo["id_grado"],
                catalogo["id_ciclo"],
            ),
        )
        notas_eliminadas = cursor.fetchall()
        conn.commit()

        return {
            "mensaje": "Curso quitado de estudiantes correctamente",
            "notas_eliminadas": len(notas_eliminadas),
        }

    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.get("/director/docentes")
def listar_docentes():
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute(
            """
            SELECT
                d.id_docente,
                d.codigo_docente,
                d.nombre_completo,
                d.estado,
                u.password_temporal,
                COALESCE(
                    string_agg(
                        DISTINCT ca.nombre || ' - ' || gr.nombre || ' - ' || cu.nombre || ' (' || ce.anio || ')',
                        ', '
                    ) FILTER (WHERE dc.id_docente_curso IS NOT NULL),
                    ''
                ) AS cursos_asignados
            FROM docentes d
            LEFT JOIN usuarios u ON u.usuario = d.codigo_docente AND u.rol = 'docente'
            LEFT JOIN docentes_cursos dc ON dc.id_docente = d.id_docente AND dc.estado = TRUE
            LEFT JOIN carreras ca ON ca.id_carrera = dc.id_carrera
            LEFT JOIN grados gr ON gr.id_grado = dc.id_grado
            LEFT JOIN cursos cu ON cu.id_curso = dc.id_curso
            LEFT JOIN ciclos_escolares ce ON ce.id_ciclo = dc.id_ciclo
            GROUP BY d.id_docente, d.codigo_docente, d.nombre_completo, d.estado, u.password_temporal
            ORDER BY d.nombre_completo;
            """
        )
        return [limpiar_fila(fila) for fila in cursor.fetchall()]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.post("/director/docentes")
def crear_docente(data: DocenteRequest):
    conn = None
    cursor = None
    try:
        password = data.password or generar_password()
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute(
            """
            INSERT INTO docentes (codigo_docente, nombre_completo, estado)
            VALUES (%s, %s, TRUE)
            ON CONFLICT (codigo_docente)
            DO UPDATE SET nombre_completo = EXCLUDED.nombre_completo, estado = TRUE
            RETURNING id_docente, codigo_docente, nombre_completo;
            """,
            (data.codigo_docente.strip(), data.nombre_completo.strip()),
        )
        docente = cursor.fetchone()
        cursor.execute(
            """
            INSERT INTO usuarios (rol, usuario, nombre_completo, password_hash, password_temporal, estado)
            VALUES ('docente', %s, %s, %s, %s, TRUE)
            ON CONFLICT (rol, usuario)
            DO UPDATE SET
                nombre_completo = EXCLUDED.nombre_completo,
                password_hash = EXCLUDED.password_hash,
                password_temporal = EXCLUDED.password_temporal,
                estado = TRUE;
            """,
            (docente["codigo_docente"], docente["nombre_completo"], hash_password(password), password),
        )
        conn.commit()
        return {**limpiar_fila(docente), "usuario": docente["codigo_docente"], "password": password}
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.post("/director/docentes/{codigo_docente}/regenerar-password")
def regenerar_password_docente(codigo_docente: str):
    return regenerar_password_usuario("docente", codigo_docente)


@app.get("/director/estudiantes")
def listar_estudiantes():
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute(
            """
            SELECT
                a.codigo_carnet,
                a.nombres,
                a.apellidos,
                a.estado,
                u.password_temporal,
                ca.nombre AS carrera,
                gr.nombre AS grado,
                ce.anio AS ciclo_escolar
            FROM alumnos a
            LEFT JOIN usuarios u ON u.usuario = a.codigo_carnet AND u.rol = 'estudiante'
            LEFT JOIN asignaciones asi ON asi.id_alumno = a.id_alumno
            LEFT JOIN carreras ca ON ca.id_carrera = asi.id_carrera
            LEFT JOIN grados gr ON gr.id_grado = asi.id_grado
            LEFT JOIN ciclos_escolares ce ON ce.id_ciclo = asi.id_ciclo
            ORDER BY ca.nombre, gr.numero, a.apellidos, a.nombres;
            """
        )
        return [limpiar_fila(fila) for fila in cursor.fetchall()]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.post("/director/estudiantes")
def crear_estudiante(data: EstudianteRequest):
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        resultado = crear_o_actualizar_estudiante(cursor, data)
        conn.commit()
        return resultado
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


def crear_o_actualizar_estudiante(cursor, data: EstudianteRequest):
    password = data.password or generar_password()
    cursor.execute(
        """
        INSERT INTO alumnos (codigo_carnet, nombres, apellidos, estado)
        VALUES (%s, %s, %s, TRUE)
        ON CONFLICT (codigo_carnet)
        DO UPDATE SET nombres = EXCLUDED.nombres, apellidos = EXCLUDED.apellidos, estado = TRUE
        RETURNING id_alumno, codigo_carnet, nombres, apellidos;
        """,
        (data.codigo_carnet.strip(), data.nombres.strip(), data.apellidos.strip()),
    )
    alumno = cursor.fetchone()

    cursor.execute(
        """
        SELECT ca.id_carrera, gr.id_grado, ce.id_ciclo
        FROM carreras ca
        CROSS JOIN grados gr
        CROSS JOIN ciclos_escolares ce
        WHERE ca.nombre = %s AND gr.nombre = %s AND ce.anio = %s;
        """,
        (data.carrera, data.grado, data.ciclo_escolar),
    )
    catalogo = cursor.fetchone()
    if not catalogo:
        raise HTTPException(status_code=404, detail="No se encontró carrera, grado o ciclo escolar")

    cursor.execute(
        """
        INSERT INTO asignaciones (id_alumno, id_carrera, id_grado, id_ciclo)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (id_alumno, id_ciclo)
        DO UPDATE SET id_carrera = EXCLUDED.id_carrera, id_grado = EXCLUDED.id_grado
        RETURNING id_asignacion;
        """,
        (alumno["id_alumno"], catalogo["id_carrera"], catalogo["id_grado"], catalogo["id_ciclo"]),
    )
    asignacion = cursor.fetchone()

    if data.curso:
        cursor.execute("SELECT id_curso FROM cursos WHERE nombre = %s;", (data.curso,))
        curso = cursor.fetchone()
        if not curso:
            raise HTTPException(status_code=404, detail="No se encontró el curso")
        cursor.execute(
            """
            INSERT INTO cursos_por_grado (id_carrera, id_grado, id_curso)
            VALUES (%s, %s, %s)
            ON CONFLICT (id_carrera, id_grado, id_curso) DO NOTHING;
            """,
            (catalogo["id_carrera"], catalogo["id_grado"], curso["id_curso"]),
        )
        cursor.execute(
            """
            INSERT INTO notas (id_asignacion, id_curso, id_bimestre, actitudinal, zona, examen, observacion)
            SELECT %s, %s, b.id_bimestre, 0, 0, 0, 'Curso asignado por dirección'
            FROM bimestres b
            ON CONFLICT (id_asignacion, id_curso, id_bimestre) DO NOTHING;
            """,
            (asignacion["id_asignacion"], curso["id_curso"]),
        )

    cursor.execute(
        """
        INSERT INTO usuarios (rol, usuario, nombre_completo, password_hash, password_temporal, estado)
        VALUES ('estudiante', %s, %s, %s, %s, TRUE)
        ON CONFLICT (rol, usuario)
        DO UPDATE SET
            nombre_completo = EXCLUDED.nombre_completo,
            password_hash = EXCLUDED.password_hash,
            password_temporal = EXCLUDED.password_temporal,
            estado = TRUE;
        """,
        (
            alumno["codigo_carnet"],
            f"{alumno['nombres']} {alumno['apellidos']}",
            hash_password(password),
            password,
        ),
    )
    return {**limpiar_fila(alumno), "usuario": alumno["codigo_carnet"], "password": password}


@app.post("/director/estudiantes/{codigo_carnet}/regenerar-password")
def regenerar_password_estudiante(codigo_carnet: str):
    return regenerar_password_usuario("estudiante", codigo_carnet)


def regenerar_password_usuario(rol, usuario):
    conn = None
    cursor = None
    try:
        password = generar_password()
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute(
            """
            UPDATE usuarios
            SET password_hash = %s,
                password_temporal = %s,
                estado = TRUE
            WHERE rol = %s AND usuario = %s
            RETURNING rol, usuario, nombre_completo, password_temporal;
            """,
            (hash_password(password), password, rol, usuario),
        )
        cuenta = cursor.fetchone()
        if not cuenta:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        conn.commit()
        return limpiar_fila(cuenta)
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.post("/director/estudiantes/importar-excel")
async def importar_alumnos_excel_director(archivo: UploadFile = File(...)):
    conn = None
    cursor = None
    try:
        if not archivo.filename.lower().endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Solo se permiten archivos .xlsx")

        contenido = await archivo.read()
        libro = load_workbook(filename=BytesIO(contenido), data_only=True)
        hoja = libro.active
        encabezados = {}
        for columna in range(1, hoja.max_column + 1):
            encabezados[normalizar_encabezado(hoja.cell(row=1, column=columna).value)] = columna

        col_codigo = encabezados.get("codigo_carnet") or encabezados.get("codigo") or encabezados.get("carnet")
        col_nombres = encabezados.get("nombres")
        col_apellidos = encabezados.get("apellidos")
        col_carrera = encabezados.get("carrera")
        col_grado = encabezados.get("grado")
        col_ciclo = encabezados.get("ciclo") or encabezados.get("ciclo_escolar")
        col_curso = encabezados.get("curso")

        if not col_codigo or not col_nombres or not col_apellidos or not col_carrera or not col_grado or not col_ciclo:
            raise HTTPException(
                status_code=400,
                detail="El Excel debe tener: codigo_carnet, nombres, apellidos, carrera, grado y ciclo",
            )

        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        procesados = 0
        resultados = []
        errores = []

        for fila in range(2, hoja.max_row + 1):
            codigo = obtener_valor_fila(hoja, fila, col_codigo)
            if codigo is None or str(codigo).strip() == "":
                continue
            try:
                data = EstudianteRequest(
                    codigo_carnet=str(codigo).strip(),
                    nombres=str(obtener_valor_fila(hoja, fila, col_nombres) or "").strip(),
                    apellidos=str(obtener_valor_fila(hoja, fila, col_apellidos) or "").strip(),
                    carrera=str(obtener_valor_fila(hoja, fila, col_carrera) or "").strip(),
                    grado=str(obtener_valor_fila(hoja, fila, col_grado) or "").strip(),
                    ciclo_escolar=int(float(obtener_valor_fila(hoja, fila, col_ciclo))),
                    curso=str(obtener_valor_fila(hoja, fila, col_curso) or "").strip() if col_curso else None,
                )
                if not data.nombres or not data.apellidos or not data.carrera or not data.grado:
                    raise ValueError("Hay campos obligatorios vacíos")
                resultado = crear_o_actualizar_estudiante(cursor, data)
                resultados.append(resultado)
                procesados += 1
            except Exception as e:
                errores.append({"fila": fila, "codigo_carnet": str(codigo), "error": str(e)})

        conn.commit()
        return {
            "mensaje": "Importación de alumnos finalizada",
            "procesados": procesados,
            "resultados": resultados,
            "errores": errores,
        }
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error al importar alumnos: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.get("/director/cursos")
def listar_cursos():
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute("SELECT id_curso, nombre FROM cursos ORDER BY nombre;")
        return [limpiar_fila(fila) for fila in cursor.fetchall()]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.post("/director/cursos")
def crear_curso(data: CursoRequest):
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute(
            """
            INSERT INTO cursos (nombre)
            VALUES (%s)
            ON CONFLICT (nombre) DO UPDATE SET nombre = EXCLUDED.nombre
            RETURNING id_curso, nombre;
            """,
            (data.nombre.strip(),),
        )
        curso = cursor.fetchone()
        if data.carrera and data.grado:
            cursor.execute("SELECT id_carrera FROM carreras WHERE nombre = %s;", (data.carrera,))
            carrera = cursor.fetchone()
            cursor.execute("SELECT id_grado FROM grados WHERE nombre = %s;", (data.grado,))
            grado = cursor.fetchone()
            if carrera and grado:
                cursor.execute(
                    """
                    INSERT INTO cursos_por_grado (id_carrera, id_grado, id_curso)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (id_carrera, id_grado, id_curso) DO NOTHING;
                    """,
                    (carrera["id_carrera"], grado["id_grado"], curso["id_curso"]),
                )
        conn.commit()
        return limpiar_fila(curso)
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.get("/director/asignaciones-docentes")
def listar_asignaciones_docentes():
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute(
            """
            SELECT
                dc.id_docente_curso,
                d.codigo_docente,
                d.nombre_completo AS docente,
                ca.nombre AS carrera,
                gr.nombre AS grado,
                cu.nombre AS curso,
                ce.anio AS ciclo_escolar,
                dc.estado
            FROM docentes_cursos dc
            INNER JOIN docentes d ON d.id_docente = dc.id_docente
            INNER JOIN carreras ca ON ca.id_carrera = dc.id_carrera
            INNER JOIN grados gr ON gr.id_grado = dc.id_grado
            INNER JOIN cursos cu ON cu.id_curso = dc.id_curso
            INNER JOIN ciclos_escolares ce ON ce.id_ciclo = dc.id_ciclo
            ORDER BY d.nombre_completo, ca.nombre, gr.numero, cu.nombre;
            """
        )
        return [limpiar_fila(fila) for fila in cursor.fetchall()]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.post("/director/asignaciones-docentes")
def asignar_curso_docente(data: DocenteCursoRequest):
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute("SELECT id_docente FROM docentes WHERE codigo_docente = %s;", (data.codigo_docente,))
        docente = cursor.fetchone()
        if not docente:
            raise HTTPException(status_code=404, detail="Docente no encontrado")
        catalogo = obtener_ids_catalogo(cursor, data.carrera, data.grado, data.curso, data.ciclo_escolar)
        cursor.execute(
            """
            INSERT INTO docentes_cursos (id_docente, id_carrera, id_grado, id_curso, id_ciclo, estado)
            VALUES (%s, %s, %s, %s, %s, TRUE)
            ON CONFLICT (id_docente, id_carrera, id_grado, id_curso, id_ciclo)
            DO UPDATE SET estado = TRUE
            RETURNING id_docente_curso;
            """,
            (
                docente["id_docente"],
                catalogo["id_carrera"],
                catalogo["id_grado"],
                catalogo["id_curso"],
                catalogo["id_ciclo"],
            ),
        )
        asignacion = cursor.fetchone()
        conn.commit()
        return {"mensaje": "Curso asignado correctamente", **limpiar_fila(asignacion)}
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.delete("/director/asignaciones-docentes/{id_docente_curso}")
def quitar_curso_docente(id_docente_curso: int):
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute(
            """
            UPDATE docentes_cursos
            SET estado = FALSE
            WHERE id_docente_curso = %s
            RETURNING id_docente_curso;
            """,
            (id_docente_curso,),
        )
        asignacion = cursor.fetchone()
        if not asignacion:
            raise HTTPException(status_code=404, detail="Asignación no encontrada")
        conn.commit()
        return {"mensaje": "Curso quitado del docente", **limpiar_fila(asignacion)}
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.get("/catalogos")
def obtener_catalogos():
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        cursor.execute("SELECT nombre, tipo FROM carreras ORDER BY nombre;")
        carreras = cursor.fetchall()
        cursor.execute("SELECT nombre, numero FROM grados ORDER BY numero;")
        grados = cursor.fetchall()
        cursor.execute("SELECT nombre FROM cursos ORDER BY nombre;")
        cursos = cursor.fetchall()
        cursor.execute("SELECT nombre, numero FROM bimestres ORDER BY numero;")
        bimestres = cursor.fetchall()
        cursor.execute("SELECT anio FROM ciclos_escolares ORDER BY anio DESC;")
        ciclos = cursor.fetchall()
        cursor.execute("SELECT codigo_docente, nombre_completo FROM docentes WHERE estado = TRUE ORDER BY nombre_completo;")
        docentes = cursor.fetchall()

        return {
            "carreras": [limpiar_fila(fila) for fila in carreras],
            "grados": [limpiar_fila(fila) for fila in grados],
            "cursos": [limpiar_fila(fila) for fila in cursos],
            "bimestres": [limpiar_fila(fila) for fila in bimestres],
            "ciclos": [limpiar_fila(fila) for fila in ciclos],
            "docentes": [limpiar_fila(fila) for fila in docentes],
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en el servidor: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)
